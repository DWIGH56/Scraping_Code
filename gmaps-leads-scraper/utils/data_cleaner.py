"""
Premium data cleaning engine for scraped Google Maps leads.

Provides:
1. Advanced Phone Cleaning — landline (021) → +6221, mobile → +628
2. Rating & Review Cleanup — fill empty rating with "-", strip review text extras
3. Address Optimization — strip keyword prefix, remove Plus Codes from address
4. Opening Hours Expansion — format as structured day-range string
5. Deduplication & Filtering — remove low-value / duplicate leads
6. XLSX Export — pandas + openpyxl formatted spreadsheet
"""

import logging
import os
import re
from datetime import datetime
from typing import Optional

import pandas as pd

from config.settings import OUTPUT_DIR

logger = logging.getLogger(__name__)


# ======================================================================
# 1. ADVANCED PHONE CLEANING
# ======================================================================

def _strip_all(raw: str) -> str:
    """Remove all non-digit characters. Preserve leading '+' for detection."""
    raw = raw.strip()
    if raw.startswith("+"):
        return "+" + re.sub(r"\D", "", raw[1:])
    return re.sub(r"\D", "", raw)


def _is_valid_phone(digits: str) -> bool:
    return len(digits.lstrip("+")) >= 8


def standardize_phone(raw: Optional[str]) -> Optional[str]:
    """
    Ultra-clean phone numbers to international format.

    Rules applied in order:
    - Strip all non-digits (spaces, dashes, brackets, dots)
    - Already +62 → keep clean
    - Already +1  → keep clean
    - Landline: starts with "(0XX)" area code → +62<areacode><number>
      e.g. (021) 1234-5678 → +622112345678  (area code 21 stripped of leading 0)
    - Mobile: starts with "08" → +628<rest>
      e.g. 0812-3456-7890 → +6281234567890
    - Starts with "0" (other) → +62<rest> strip leading 0
    - Starts with "8" (Indonesian mobile without 0) → +62<rest>
    - 10-15 digit foreign → +<rest> (US default fallback)
    """
    if not raw or not raw.strip():
        return None

    digits = _strip_all(raw)

    if not _is_valid_phone(digits):
        logger.debug(f"Dropping invalid phone: {raw!r} → {digits!r}")
        return None

    stripped = digits.lstrip("+")

    # --- Already international ---
    if digits.startswith("+62"):
        return "+62" + stripped[2:]
    if digits.startswith("+1"):
        return "+1" + stripped[1:]

    # --- Landline area code detection: (021), (022), (031), etc. ---
    # Raw input like "(021) 1234" or "021 1234" indicates landline
    raw_normalised = re.sub(r"[^\d]", "", raw)  # digits only, no +
    if raw_normalised.startswith("0"):
        # Could be landline 021xxxx or mobile 0812xxxx
        after_zero = raw_normalised[1:]  # e.g. "2112345678" or "81234567890"
        if len(after_zero) >= 12:
            # Looks like mobile: 0212xxxxxxxx (long) → treat as mobile with 0 prefix
            return f"+62{after_zero}"
        elif after_zero.startswith("8"):
            # 08xxx → mobile
            return f"+62{after_zero}"
        else:
            # (021)xxxx → landline with area code
            return f"+62{after_zero}"
    # --- Starts with "8" (Indonesian mobile without 0) ---
    if stripped.startswith("8") and len(stripped) >= 9:
        return f"+62{stripped}"
    # --- Fallback: 10-15 digit → US default ---
    if 10 <= len(stripped) <= 15:
        return f"+{stripped}"
    return f"+{stripped}"


# ======================================================================
# 2. RATING & REVIEW CLEANUP
# ======================================================================

def clean_rating(rating: Optional[float]) -> str:
    """If rating is None or empty, return '-'."""
    if rating is None:
        return "-"
    return str(rating)


def clean_reviews_count(reviews_count: Optional[int]) -> int:
    """Ensure reviews_count is a pure integer; 0 if missing."""
    if reviews_count is None:
        return 0
    try:
        return int(reviews_count)
    except (ValueError, TypeError):
        return 0


# ======================================================================
# 3. ADDRESS OPTIMIZATION
# ======================================================================

# Regex to detect Plus Codes: alphanumeric pattern like "7JVW+PQ Jakarta"
PLUS_CODE_PATTERN = re.compile(
    r"\b[A-Z0-9]{4,6}\+[A-Z0-9]{2,4}\b", re.IGNORECASE
)


def clean_address(address: Optional[str], keyword: Optional[str] = None) -> Optional[str]:
    """
    Clean a Google Maps address string.

    Steps:
    1. Strip the search keyword prefix if it appears at the start
       (e.g. "Plumber · Jl. Sudirman" → "Jl. Sudirman")
    2. Remove Plus Codes (alphanumeric short codes like "7JVW+PQ")
    3. Title-case the result
    """
    if not address or not address.strip():
        return address

    cleaned = address.strip()

    # 1. Strip keyword prefix separated by " · " or " - "
    if keyword:
        # Split on common separators: " · ", " - ", " | "
        for sep in [" · ", " - ", " | "]:
            if sep in cleaned:
                parts = [p.strip() for p in cleaned.split(sep)]
                # If the first part looks like the keyword (case-insensitive)
                if keyword.lower() in parts[0].lower():
                    cleaned = " ".join(parts[1:])
                    break

    # 2. Remove Plus Codes from the address string
    #    e.g. "Jl. Sudirman, 7JVW+PQ Jakarta" → "Jl. Sudirman, Jakarta"
    cleaned = PLUS_CODE_PATTERN.sub("", cleaned).strip()
    # Clean up double spaces / trailing commas from the removal
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    cleaned = re.sub(r",\s*,", ",", cleaned)
    cleaned = cleaned.strip(", ").strip()

    # 3. Title-case
    return _to_title_case(cleaned) if cleaned else address


# ======================================================================
# 4. OPENING HOURS EXPANSION
# ======================================================================

def clean_hours(hours_raw: Optional[str]) -> Optional[str]:
    """
    Parse raw opening hours string into a structured day-range format.

    Input Google Maps pattern variations:
      "08:00–20:00"           (single entry — applies to all days)
      "Mon-Fri: 09:00–18:00"
      "Senin-Jumat: 08:00-20:00 | Sabtu-Minggu: 09:00-17:00"
      "Mon-Sat: 09:00-21:00, Sun: 10:00-18:00"

    Output structured format (Indonesian day labels):
      "Senin-Jumat: 08:00-20:00 | Sabtu-Minggu: 09:00-17:00"

    Falls back to the raw string if parsing fails.
    """
    if not hours_raw or not hours_raw.strip():
        return None

    text = hours_raw.strip()

    # Day name mappings (English → Indonesian)
    day_map = {
        "mon": "senin", "tue": "selasa", "wed": "rabu",
        "thu": "kamis", "fri": "jumat",
        "sat": "sabtu", "sun": "minggu",
    }

    # Replace English day abbreviations with Indonesian
    for en, idn in day_map.items():
        text = re.sub(rf"\b{en}\b", idn, text, flags=re.IGNORECASE)

    # Capitalise Indonesian day names
    def _cap_day(m):
        return m.group(0).capitalize()
    for idn in day_map.values():
        text = re.sub(rf"\b{idn}\b", _cap_day, text, flags=re.IGNORECASE)

    # If the result already contains "|" or " - " separator, it's already structured
    if "|" in text or " - " in text or "–" in text or "-" in text:
        # Normalise dashes
        text = text.replace("–", "-").replace("—", "-")
        return text

    # Single "HH:MM-HH:MM" without day labels → wrap with "All days"
    if re.match(r"^\d{2}:\d{2}\s*[-–]\s*\d{2}:\d{2}", text):
        return f"Senin-Minggu: {text}"

    return text


# ======================================================================
# TEXT FORMATTING HELPERS
# ======================================================================

def _to_title_case(text: Optional[str]) -> Optional[str]:
    if not text or not text.strip():
        return text
    words = text.strip().split()
    return " ".join(w.capitalize() for w in words)


def _clean_website_url(url: Optional[str]) -> Optional[str]:
    if not url or not url.strip():
        return url
    cleaned = url.strip()
    if "?" in cleaned:
        cleaned = cleaned.split("?")[0].rstrip("/")
    if "#" in cleaned:
        cleaned = cleaned.split("#")[0].rstrip("/")
    return cleaned if cleaned else url.strip()


# ======================================================================
# DEDUPLICATION & FILTERING
# ======================================================================

def filter_low_value(leads: list[dict]) -> list[dict]:
    """Remove leads with NEITHER phone NOR website."""
    clean: list[dict] = []
    for lead in leads:
        phone = (lead.get("phone") or "").strip()
        website = (lead.get("website") or "").strip()
        if phone or website:
            clean.append(lead)
    filtered = len(leads) - len(clean)
    if filtered:
        logger.info(f"Filtered out {filtered} low-value lead(s) (no phone & no website).")
    return clean


def deduplicate(leads: list[dict]) -> list[dict]:
    seen_phones: set[str] = set()
    seen_names: set[str] = set()
    clean: list[dict] = []
    for lead in leads:
        phone = (lead.get("phone") or "").strip()
        name = (lead.get("name") or "").strip().lower()
        if phone and phone in seen_phones:
            continue
        if not phone and name and name in seen_names:
            continue
        if phone:
            seen_phones.add(phone)
        if name:
            seen_names.add(name)
        clean.append(lead)
    duplicates_removed = len(leads) - len(clean)
    if duplicates_removed:
        logger.info(f"Deduplication removed {duplicates_removed} duplicate(s).")
    return clean


def _format_name(name: Optional[str]) -> Optional[str]:
    return _to_title_case(name) if name else name


# ======================================================================
# 5. FULL CLEANING PIPELINE
# ======================================================================

def clean_leads(leads: list[dict]) -> list[dict]:
    """
    Run the full premium cleaning pipeline.

    Steps:
    1. Standardise phone numbers (landline + mobile → +62)
    2. Clean website URLs (remove tracking)
    3. Clean addresses (remove Plus Codes, keyword prefix, title-case)
    4. Clean ratings (empty → "-")
    5. Clean reviews count (ensure pure int, 0 if missing)
    6. Clean / expand opening hours
    7. Title-case business names
    8. Filter low-value leads (no phone + no website)
    9. Deduplicate by phone / name
    """
    if not leads:
        return leads

    logger.info(f"🧼 Premium cleaning {len(leads)} leads...")

    for lead in leads:
        # 1. Phone
        if lead.get("phone"):
            lead["phone"] = standardize_phone(lead["phone"])

        # 2. Website
        if lead.get("website"):
            lead["website"] = _clean_website_url(lead["website"])

        # 3. Address — strip Plus Codes + keyword prefix
        if lead.get("address"):
            keyword = lead.get("keyword")
            lead["address"] = clean_address(lead["address"], keyword=keyword)

        # 4. Rating — empty → "-"
        lead["rating"] = clean_rating(lead.get("rating"))

        # 5. Reviews count — ensure pure int
        lead["reviews_count"] = clean_reviews_count(lead.get("reviews_count"))

        # 6. Opening hours — expand / structure
        if lead.get("hours"):
            lead["hours"] = clean_hours(lead["hours"])

        # 7. Title-case name
        if lead.get("name"):
            lead["name"] = _format_name(lead["name"])

    # 8. Filter low-value
    leads = filter_low_value(leads)

    # 9. Deduplicate
    leads = deduplicate(leads)

    # 10. Final dash-rule enforcement: every field that is None/empty/0 → "-"
    for lead in leads:
        for key in list(lead.keys()):
            lead[key] = _apply_dash_rule(lead[key], key)

    logger.info(f"✅ Premium cleaning complete: {len(leads)} leads remaining.")
    return leads


# ======================================================================
# 6. XLSX EXPORT (Formatted Excel)
# ======================================================================

XLSX_FILENAME = "gmaps_leads.xlsx"

# A dash replaces any missing/null/empty/zero value
_DASH = "-"

# All string-type columns that must be "-" if missing
_STRING_COLS = {"phone", "website", "instagram", "facebook", "tiktok", "address", "hours", "plus_code", "keyword", "location"}


def _apply_dash_rule(value, col_name: str):
    """
    Enforce the strict dash rule:
    - If value is None, empty string, 0, or 0.0 → return "-"
    - Otherwise return the value as-is
    """
    if value is None or value == "" or value == 0 or value == 0.0:
        return _DASH
    return value


# 12-column premium schema — reordered: Name, Address, Phone, Email, Social, Website, Ads, Rating
SPREADSHEET_COLUMNS = [
    "no",
    "name",
    "address",
    "phone",
    "email",
    "instagram",
    "tiktok",
    "facebook",
    "website",
    "punya_ads",
    "rating",
    "reviews_count",
]

HEADER_LABELS: dict[str, str] = {
    "no": "No",
    "name": "Name",
    "address": "Address",
    "phone": "Phone",
    "email": "Email",
    "instagram": "Instagram",
    "tiktok": "TikTok",
    "facebook": "Facebook",
    "website": "Website",
    "punya_ads": "Punya Ads",
    "rating": "Rating",
    "reviews_count": "Review Count",
}


def _prepare_dataframe(leads: list[dict]) -> pd.DataFrame:
    rows = []
    for idx, lead in enumerate(leads, start=1):
        row = {}
        # Compute has_ads from both ad flags
        has_ads = bool(lead.get("has_google_ads") or lead.get("has_facebook_pixel"))
        for col in SPREADSHEET_COLUMNS:
            if col == "no":
                val = idx
            elif col == "punya_ads":
                val = "✅ Yes" if has_ads else "❌ No"
            else:
                val = lead.get(col)
                val = _apply_dash_rule(val, col)
            row[HEADER_LABELS.get(col, col)] = val
        rows.append(row)
    return pd.DataFrame(rows)

BOOL_MAP = {True: "✅ Yes", False: "❌ No", None: _DASH}


def export_to_xlsx(
    leads: list[dict],
    filename: str = XLSX_FILENAME,
    sheet_name: str = "Leads",
) -> str:
    """
    Export leads to a professionally formatted .xlsx file.

    Features:
    - Bold, coloured headers (Google Blue #1A73E8)
    - Auto-adjusted column widths (clamped 10–50)
    - Thin grey gridlines
    - Freeze top row
    - Auto-filter on all columns
    - Column order optimised for lead buyers
    """
    filepath = os.path.join(OUTPUT_DIR, filename)

    if not leads:
        logger.warning("No leads to export to XLSX.")
        return filepath

    df = _prepare_dataframe(leads)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="top")

        for col_idx, col_cells in enumerate(worksheet.columns, start=1):
            cell = col_cells[0]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        # Auto-fit column widths
        for col_cells in worksheet.columns:
            col_letter = col_cells[0].column_letter
            lengths = []
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    lengths.append(len(val))
                except Exception:
                    pass
            best = max(lengths) if lengths else 10
            worksheet.column_dimensions[col_letter].width = max(min(best + 3, 50), 10)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    logger.info(f"✅ XLSX exported: {filepath}")
    return filepath